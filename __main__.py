# -*- coding: utf-8 -*-
"""
2022/04/19
original
H. Uryu, 1521512@ed.tus.ac.jp(2023卒)

2022/4/28
Pulse化
R. Kaneko 1519032@ed.tus.ac.jp(2025卒)
K. Tomiyoshi 1522529@ed.tus.ac.jp(2024卒)

2022/5/1
GUI化、ファイル出力、強制終了ボタン、各種選択機能の追加
M. Taniguchi 1521536@ed.tus.ac.jp(2023卒)
"""
#default設定
interval_time = 0.041463354054055365#[s] 実行環境によって異なるので適時調整
d_V_bot = 0.1#[V]
d_bot_time = 10#[s]
d_V_top = 0.8#[V]
d_top_time = 3.0#[s]
d_loop = 2#回
d_hip = 0#[s]
d_folderpath = 'C:/Users/higuchi/Desktop/パルス測定'
d_l_interval = 0.2#[s]ライブ描画の更新間隔
d_x_label = "Time [s]"
d_y1_label = "Voltage [V]"
d_y2_label = "Current [A]"

import matplotlib.pyplot as plt
import os
import pyvisa as visa
import threading
import time
import tkinter as tk
from tkinter import filedialog
from tkinter import ttk
 
class debug:
    def __init__(self, list1):
        self._x = list1
        
    #縦軸のみのグラフで表示、バラつき見る用 
    def dispersion(self):
        x = [1 for _ in range(len(self._x))]  
        fig = plt.figure()
        ax1 = fig.add_subplot()
        ax1.scatter(x, self._x)
        ax1.set_ylabel("Time [s]")
        plt.show()
        
    #平均値を取得 
    def mean(self):
        print(sum(self._x)/len(self._x))

rm = visa.ResourceManager(r'C:\WINDOWS\system32\visa64.dll')
dev = rm.open_resource('GPIB1::1::INSTR')
dev.timeout = 5000

#送信コマンド
def write(command):
    dev.write(command)

print(dev.query('*IDN?'))

#フォルダ選択
def set_folder_func():
    dir = 'C:\\'
    folder_path = filedialog.askdirectory(initialdir = dir)
    textbox["folderpath"].delete(0, tk.END)
    textbox["folderpath"].insert(tk.END, folder_path)       

#グラフ
def graph(x_list, y1_list, y2_list, plot, scatter):
    def para(dic):
        return {f'{k1}.{k2}' : v for k1, d in dic.items() for k2, v in d.items()} 
    config = {
        "font" :{
            "family":"Times New Roman",
            "size":14
            },
        "xtick" :{
                "direction":"in",
                "top":True,
                "major.width":1.2,
                "labelsize":20.0
            },
        "ytick" :{
                "direction":"in",
                "right":True,
                "major.width":1.2,
                "labelsize":20.0
            },
        "axes" :{
            "linewidth":1.2,
            "labelpad":10
            },
        
        "figure" :{
            "dpi":150
                }
        }
    
    plt.rcParams.update(para(config))
    
    fig=plt.figure()
    ax1=fig.add_subplot(2, 1, 1)
    ax2=fig.add_subplot(2, 1, 2)
    
    if plot == True:
        ax1.plot(x_list, y1_list)
        ax2.plot(x_list, y2_list)
    if scatter == True:
        ax1.scatter(x_list, y1_list)
        ax2.scatter(x_list, y2_list)
    
    ax1.set_xlabel(d_x_label)
    ax1.set_ylabel(d_y1_label)
    ax2.set_xlabel(d_x_label)
    ax2.set_ylabel(d_y2_label)    
    plt.show()

#強制停止
def stop_func():
    global stop_flag
    stop_flag = True
    swrite("測定中断")

#測定タイマー
def timer(measure_times):
    global time_list, timer_flag
    start_time = time.perf_counter()
    while 1:
        if timer_flag == True:
            swrite(f"合計時間: {time.perf_counter()-start_time:.1f} [s]")
            break
        swrite(f"経過時間: {time.perf_counter()-start_time:.1f} [s]" + "," + f"{len(time_list)}/" + str(measure_times))
        time.sleep(0.1)

#ライブ描画
def livegraph(plot, scatter):
    global livegraph_flag, time_list, A_list, V_list

    time.sleep(1.0)
    while 1:
        if livegraph_flag == True:
            break
        time_list_ = time_list

        V_list_ = [V_list[i] for i in range(len(time_list_)-1)] 
        A_list_ = [A_list[i] for i in range(len(time_list_)-1)]
        interval_list = [time_list[i+1]-time_list[i] for i in range(len(time_list)-2)]
        totaltime_list = [sum(interval_list[:i]) for i in range(len(interval_list)+1)]
        
        time.sleep(d_l_interval)
        
        graph(totaltime_list, V_list_, A_list_, True, False)
        
#測定
def measure(V_set, measure_times, interval_time):
    global stop_flag, time_list
    for _ in range(measure_times):
        if stop_flag == True:
            break    
        
        write(f"SOV{V_set}")
        write("*TRG")
        time_list.append(time.perf_counter())
        
        A=dev.query("N?")
        A_=float(A[3:-2])
        A_list.append(A_)
        
        V=dev.query("SOV?")    
        V_=float(V[3:-2])
        V_list.append(V_)

#ファイル出力
def output(filepath, list1, list2, list3, extension_index):
    def output_txt():
        with open(filepath, 'w') as data:
            for list1_, list2_, list3_ in zip(list1, list2, list3):
                data.write(f"{str(list1_)} {str(list2_)} {str(list3_)}\n")    
    
    def output_csv():
        import csv
        with open(filepath, 'w', newline="") as data:
            writer = csv.writer(data)
            for list1_, list2_, list3_ in zip(list1, list2, list3):
                writer.writerow([list1_, list2_, list3_])
                
    def output_xlsx():
        from openpyxl import Workbook
        from openpyxl import load_workbook
        
        wb = Workbook()
        wb.save(filepath)
        wb = load_workbook(filepath)
        ws = wb['Sheet']
        ws = wb.active
        
        for i, (t, V_val, A_val) in enumerate(zip(list1, list2, list3), 1):
            ws.cell(i, 1, t)
            ws.cell(i, 2, V_val)
            ws.cell(i, 3, A_val)
            
        wb.save(filepath)
        wb.close()    
                
    if extension_index == 0:
        output_txt()          
    if extension_index == 1:
        output_csv()
    if extension_index == 2:
        output_xlsx()

#実行
def run_func():
    global A_list, V_list, time_list, stop_flag, timer_flag, livegraph_flag
    global interval_list#debug用
    stop_flag = False
    V_list, A_list, time_list =[], [], []
    
    #値を取得
    loop = float(spinbox["ループ回数"].get())
    V_bot = float(spinbox["V_bot"].get())
    V_top = float(spinbox["V_top"].get())
    top_time = float(spinbox["top_time"].get())#[s]
    bot_time = float(spinbox["bot_time"].get())#[s]
    hip_time = float(spinbox["おしり"].get())
    
    chk0 = checkbutton["ファイルに出力しない"].get()
    chk1 = checkbutton["測定終了後、プロットを表示する"].get()
    chk2 = checkbutton["測定終了後、散布図を表示する"].get()
    chk3 = checkbutton["タイマーを無効にする"].get()
    chk4 = checkbutton["ライブ描画を有効にする"].get()
    extension_box_index = combobox["ext"].current()
    extension = combobox["ext"].get()
    
    bot_times = int(bot_time/2/interval_time)#切り捨て
    top_times = int(top_time/interval_time)#切り捨て
    hip_times = int(hip_time/interval_time) - bot_times#切り捨て
    if hip_times > 0: 
        measuretimes = (bot_times*2+top_times)*int(loop) + hip_times
    else:
        measuretimes = (bot_times*2+top_times)*int(loop)
    
    #エラーチェック
    if not chk0 == True:
        folderpath = textbox["folderpath"].get()
        filename = textbox["filename"].get()
        if not os.path.exists(folderpath):
            swrite("無効なフォルダーパスです")
            return
        filepath = folderpath +'/' + filename + extension
        
    if bot_times < interval_time:
        swrite("bot_timeが短すぎます")
        return
    
    if top_times < interval_time:
        swrite("top_timeが短すぎます")
        return
    
    if loop.is_integer() == False:
        swrite("ループ回数は整数値を設定して下さい")
        return
    
    #測定装置の準備
    write("*RST")#初期化
    write("M1")#トリガーモード HOLD
    write("OH1")#ヘッダON
    write("VF")#電圧発生
    write("F2")#電流測定
    write("MD0")#DCモード
    write("R0")#オートレンジ
    write("OPR")#出力
    
    #測定タイマー起動
    if chk3 == False:
        timer_flag = False
        t1 = threading.Thread(target = timer, args = (measuretimes,))
        t1.start()
    else:
        swrite("測定中")
    #ライブ描画起動
    if chk4 == True:
        livegraph_flag = False
        t2 = threading.Thread(target = livegraph, args = (chk1, chk2))
        t2.start()
    #測定実行
    for _ in range(int(loop)):
        measure(V_bot, bot_times, interval_time)         
        measure(V_top, top_times, interval_time)
        measure(V_bot, bot_times, interval_time)
    if hip_times > 0:
        measure(V_bot, hip_times, interval_time)
    
    #タイマー終了
    if chk3 == False:
        timer_flag = True  
    else:
        swrite("測定終了")
    #ライブ描画終了
    if chk4 == True:
        livegraph_flag = True
    
    write("SBY")
    
    #時間軸の作成
    interval_list = [time_list[i+1]-time_list[i] for i in range(len(time_list)-1)]
    totaltime_list = [sum(interval_list[:i]) for i in range(len(interval_list)+1)]

    graph(totaltime_list, A_list, V_list, chk1, chk2)
    
    #ファイルに出力する場合
    if not chk0 == True:
        output(filepath, totaltime_list, V_list, A_list, extension_box_index)

def exc_run_func():
    try:      
        t = threading.Thread(target = run_func)
        t.start()

    except:
        swrite("予期せぬエラーです")

root = tk.Tk()
root.title("Pulse ver1.1")
root.geometry("430x300")#横×縦
root.resizable(False, False)#ウィンドウサイズをフリーズ
root.lift()#最前面に表示

#ラベル
def create_label(config):
             for var in config: 
                if var[5] == True:
                    label[var[0]] = tk.Label(text= var[0], background= '#B0E0E6')
                else:
                    label[var[0]] = tk.Label(text= var[0])
                label[var[0]].place(x=var[1] + var[2]*var[6], y= var[3] + var[4]*var[6])

label = {} 
label_list = [['保存先のフォルダ', 'ファイル名を入力'],
              ['V_top [V]', 'top_time [s]', 'V_bot [V]', 'bot_time [s]', 'ループ回数', 'おしり [s]'],
              ['※有効の場合、若干ばらつきが増加'],
              ['ファイル形式'],]              
#x = a+bx, y=c+dxを満たす[a, b, c, d] + background   
label_params = [[25, 0, 10, 30, True],
                [40, 0, 75, 25, False],
                [230, 0, 172, 0, False],
                [290, 0, 40, 0, True],]
label_config = [[tag_] + con + [i] for tag, con in zip(label_list, label_params) for i, tag_ in enumerate(tag)]
create_label(label_config)

#テキストボックス
def create_textbox(config):
    for key, var in config.items():
        textbox[key] = ttk.Entry(width= var[0])
        textbox[key].place(x= var[1], y= var[2])
        textbox[key].insert(0, var[3])
        
textbox = {}
textbox_config = {
    #{tag :[wid, x, y, init]}
    "folderpath" :[38, 120, 10, d_folderpath],
    "filename" :[25, 120, 40, ""],
    }  
create_textbox(textbox_config)

#スピンボックス
def create_spinbox(config):
    for i, (key, var) in enumerate(config.items()):
        spinbox[key] = ttk.Spinbox(
            root, 
            width = 7,
            format = '%3.1f',
            from_ = var[0],
            to = var[1],
            increment = var[2],
            )            
        spinbox[key].place(x= 125, y= 75 + 25*i)
        spinbox[key].insert(0, var[3])

spinbox = {}
spinbox_config = {
    #{tag :[min, max, step, init]}
    "V_top" :[-30.0, 30.0, 0.1, d_V_top],
    "top_time" :[-30.0, 30.0, 0.1, d_top_time],
    "V_bot" :[-30.0, 30.0, 0.1, d_V_bot],
    "bot_time" :[0.0, 10000.0, 0.1, d_bot_time],
    "ループ回数":[1, 10000, 1, d_loop],
    "おしり":[1, 10000, 0, d_hip],
    }
create_spinbox(spinbox_config)

#ボタン
def create_button(config):
    for key, var in config.items():
        button[key] = ttk.Button(
            root,
            text = key,
            width = var[0],
            padding = [var[1], var[2]],
            command = var[5],
            )
        button[key].place(x= var[3], y= var[4])
        
button = {}
button_config = {
    #{tag :[wid, pad_EW, pad_NS, x, y, command]}
    "参照": [8, 0, 0, 360, 9, set_folder_func],
    "実行": [12, 0, 10, 125, 225, exc_run_func],
    "強制終了": [12, 0, 10, 225, 225, stop_func],
    }
create_button(button_config)

#チェックボタン
def create_checkbutton(config):
    for i, (key, var) in enumerate(config.items()):
        checkbutton[key] = tk.BooleanVar()
        checkbutton[key].set(var)
        chk = ttk.Checkbutton(
            root,
            variable = checkbutton[key],
            text = key
            )
        chk.place(x= 230, y= 75 + 20*i)

checkbutton = {}
checkbutton_config = {
    #[text :bln]
    'ファイルに出力しない' :False,
    '測定終了後、プロットを表示する' :True,
    '測定終了後、散布図を表示する' :False,
    'タイマーを無効にする' :False,
    'ライブ描画を有効にする' :False,
    }        
create_checkbutton(checkbutton_config)

#プルダウンリスト
def create_combobox(config):
    for key, var in config.items():
        combobox[key] = ttk.Combobox(
            root,
            width = var[0],
            justify = "left", 
            state = "readonly",
            values = var[1],
            )
        combobox[key].place(x= var[2], y= var[3])
        combobox[key].current(var[4])
        
combobox = {}
combobox_config = {
    #tag :[wid, [values], x, y, init]
    "ext": [4, [".txt", ".csv", ".xlsx"], 360, 40, 2],       
    }
create_combobox(combobox_config)

statusbar = tk.Label(root, text = "", bd = 1, relief = tk.SUNKEN, anchor = tk.W)
statusbar.pack(side = tk.BOTTOM, fill = tk.X)
def swrite(text):
    statusbar["text"] = text

root.mainloop()