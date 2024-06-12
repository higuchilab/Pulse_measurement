class Datas():
  def __init__(self):
    self.time_list = []
    self.A_list = []
    self.V_list = []
    self.timer_flag = False

  def output(self, filepath, extension_index):
    if extension_index == 0:
      with open(filepath, 'w') as data:
        for time, voltage, current in zip(self.time_list, self.V_list, self.A_list):
          data.write(f"{str(time)} {str(voltage)} {str(current)}\n")  

    if extension_index == 1:
      import csv
      with open(filepath, 'w', newline="") as data:
        writer = csv.writer(data)
        for time, voltage, current in zip(self.time_list, self.V_list, self.A_list):
          writer.writerow([time, voltage, current])

    if extension_index == 2:
      from openpyxl import Workbook
      from openpyxl import load_workbook
      
      wb = Workbook()
      wb.save(filepath)
      wb = load_workbook(filepath)
      ws = wb['Sheet']
      ws = wb.active
      
      for i, (t, V_val, A_val) in enumerate(zip(self.time_list, self.V_list, self.A_list), 1):
        ws.cell(i, 1, t)
        ws.cell(i, 2, V_val)
        ws.cell(i, 3, A_val)
          
      wb.save(filepath)
      wb.close()    
