import time
import os
import numpy as np
from numpy.typing import NDArray

from dotenv import load_dotenv
from typing import TypedDict, Protocol, Any
from abc import ABC, abstractmethod
from enum import Enum, auto
from openpyxl import Workbook, load_workbook

from ..visualization import graph
from .data_processing import TwoTerminalOutput, EchoStateOutput, NarmaParam, SweepParam, CommonParameters, HistoryParam
from .measurement_model import MeasureBlocks, PulseModel, MeasureModelTemplete, SweepModel
from ..utils import plot_data
from .device_control import write_command, prepare_device, device_connection

from .measurement_strategies import (
    MeasurementStrategy
)


# 定数の分離
# class Constants:
#     VISA_DLL_PATH = r'C:\WINDOWS\system32\visa64.dll'
#     GPIB_ADDRESS = 'GPIB1::1::INSTR'
#     INTERVAL_TIME = 0.041463354054055365

# 測定タイプの列挙
class MeasurementType(Enum):
    PULSE = "2-terminal Pulse"
    NARMA = "NARMA"
    SWEEP = "2-terminal I-Vsweep"


def stop_func(statusbar: Any) -> None:
    """測定を中断し、ステータスバーに表示します。"""
    statusbar.swrite("測定中断")


# メイン測定クラス
class MeasurementExecutor:
    def __init__(self, strategy: MeasurementStrategy, common_param: CommonParameters):
        self.strategy = strategy
        self.common_param = common_param
        self.device = None

    def _connect_device(self):
        """デバイスへの接続"""
        load_dotenv()
        gpib_address = os.environ['GPIB_ADDRESS_TWO_TERMINAl']
        try:
            self.device = device_connection(
                visa_dll_path=r'C:\WINDOWS\system32\visa64.dll',
                gpib_address=gpib_address
            )
            prepare_device(self.device)
        except Exception as e:
            raise ConnectionError(f"Failed to connect device '{gpib_address}': {str(e)}")

    def _save_results(self, output: TwoTerminalOutput, header: list[str] | None) -> None:
        """
        結果の保存
        file_pathのextensionによって保存形式を変更する
        """
        if self.common_param.file_path.endswith(".xlsx"):
            print("測定結果をExcelに保存中")
            output_to_excel_file(file_path=self.common_param.file_path, output=output)
            print("測定結果を保存しました")

        elif self.common_param.file_path.endswith(".csv"):
            print("測定結果をCSVに保存中")
            output_to_csv(file_path=self.common_param.file_path, output=output, header=header)
            print("測定結果を保存しました")

    def execute(self) -> TwoTerminalOutput:
        """測定の実行"""
        try:
            self._connect_device()
            measure_model = self.strategy.create_measure_model()
            print(f"測定モデル: {measure_model}")

            print("測定開始")
            output = self.strategy.measure(measure_model=measure_model, dev=self.device)
            write_command("SBY", self.device)
            print("測定終了")

            output = self.strategy.data_formatting(output)
            self.strategy.save_to_db(self.common_param, output)
            header = self.strategy.get_header()
            self.strategy.post_process(output)
            self._save_results(output, header)

            return output

        except Exception as e:
            print(f"測定中にエラーが発生しました: {str(e)}")
            raise e

        finally:
            if self.device:
                write_command("SBY", self.device)


def output_to_excel_file(file_path: str, output: TwoTerminalOutput):
    wb = Workbook()
    wb.save(file_path)
    wb = load_workbook(file_path)
    ws =wb['Sheet']
    ws = wb.active

    ws['A1'] = "Time"
    ws['B1'] = "Voltage"
    ws['C1'] = "Current"

    for i, (t, voltage, current) in enumerate(zip(output.time, output.voltage, output.current), 2):
        ws.cell(i, 1, t)
        ws.cell(i, 2, voltage)
        ws.cell(i, 3, current)

    wb.save(file_path)
    wb.close()


def output_to_csv(file_path: str, output: NDArray, header: list[str] | None = None):
    """
    測定結果をcsvファイルに保存する関数
    headerを指定しない場合は、ヘッダー行は保存されません。
    1行目にヘッダー行を追加する場合は、headerにリストを指定してください。
    例: header=["Time", "Voltage", "Current"]

    Parameters:
        file_path (str): 保存先のファイルパス
        output (NDArray): 測定結果の配列
        header (list[str], optional): ヘッダー行のリスト。デフォルトはNone。
    """
    if header is not None:
        header_str = ",".join(header)
        np.savetxt(file_path, output, delimiter=",", header=header_str, comments="")
    else:
        np.savetxt(file_path, output, delimiter=",")
