from dataclasses import dataclass
from typing import Any, TypedDict
from typing_extensions import Literal
from numpy.typing import NDArray
from pydantic import BaseModel
from pydantic.fields import Field

from openpyxl import Workbook, load_workbook

class Datas():
    def __init__(self, time_list: NDArray, current_list: NDArray, voltage_list: NDArray):
        self.__time_list = time_list
        self.__current_list = current_list
        self.__voltage_list = voltage_list

    @property
    def time_list(self):
        return self.__time_list
    
    @property
    def current_list(self):
        return self.__current_list
    
    @property
    def voltage_list(self):
        return self.__voltage_list
    
    def output(self, filepath):
        folderpath = 'C:Users/higuchi/Desktop/パルス測定'
        wb = Workbook()
        wb.save(filepath)
        wb = load_workbook(filepath)
        ws =wb['Sheet']
        ws = wb.active

        for i, (t, voltage, current) in enumerate(zip(self.time_list, self.voltage_list, self.current_list), 1):
            ws.cell(i, 1, t)
            ws.cell(i, 2, voltage)
            ws.cell(i, 3, current)

        wb.save(filepath)
        wb.close()


class CommonParameters(BaseModel):
    operator: str
    material: str
    sample_name: str
    option: str
    file_path: str


@dataclass(frozen=True)
class TwoTerminalOutput:
    voltage: NDArray
    current: NDArray
    time: NDArray


class EchoStateOutput(BaseModel):
    """
    Output data for Echo State Network
    """
    voltage: Any
    current: Any
    time: Any
    descrete_time: Any
    internal_loop: Any
    external_loop: Any


class PulseBlockParam(BaseModel):
    top_voltage: float
    top_time: float
    base_voltage: float
    base_time: float
    loop: int
    interval_time: float


class SweepParam(BaseModel):
    # mode: Literal["one_way", "round_trip", "bidirection"]
    mode: str
    top_voltage: float
    bottom_voltage: float
    voltage_step: float
    loop: int
    tick_time: float


class NarmaParam(BaseModel):
    """
    Parameters for NARMA measurement from GUI\n
    parameters:
    - use_database: bool
    - model: str
    - pulse_width: float
    - off_width: float
    - tick: float
    - nodes: int
    - discrete_time: int
    - bot_voltage: float
    - top_voltage: float
    - base_voltage: float
    """
    use_database: bool
    model: str
    pulse_width: float
    off_width: float
    tick: float
    nodes: int
    discrete_time: int
    bot_voltage: float
    top_voltage: float
    base_voltage: float


class EchoStateParam(BaseModel):
    """
    Parameters for Echo State measurement from GUI\n
    parameters:
    - pulse_width: float
    - duty_rate: float
    - tick: float
    - discrete_time: int
    - top_voltage: float
    - base_voltage: float
    - inner_loop_num: int
    - outer_loop_num: int
    """
    pulse_width: float = Field(default=1., gt=0, description="Pulse width")
    duty_rate: float = Field(default=0.5, gt=0, lt=1, description="Duty rate")
    tick: float = Field(default=0.5, gt=0, description="Tick time")
    discrete_time: int = Field(default=100, gt=0, description="Discrete time")
    top_voltage: float = Field(default=0.8, gt=-30, lt=30, description="Top voltage")
    base_voltage: float = Field(default=0., gt=-30, lt=30, description="Base voltage")
    inner_loop_num: int = Field(default=30, gt=0, description="Inner loop number")
    outer_loop_num: int = Field(default=10, gt=0, description="Outer loop number")

    def update(
            self,
            target: Literal[
                "pulse_width",
                "duty_rate",
                "tick",
                "discrete_time",
                "top_voltage",
                "base_voltage",
                "inner_loop_num",
                "outer_loop_num"
            ], 
            value
        ) -> None:
        """
        Update the parameter value
        """
        try:
            if target == "pulse_width":
                self.pulse_width = value
            elif target == "duty_rate":
                self.duty_rate = value
            elif target == "tick":
                self.tick = value
            elif target == "discrete_time":
                self.discrete_time = value
            elif target == "top_voltage":
                self.top_voltage = value
            elif target == "base_voltage":
                self.base_voltage = value
            elif target == "inner_loop_num":
                self.inner_loop_num = value
            elif target == "outer_loop_num":
                self.outer_loop_num = value
        except ValueError as e:
            print(f"Invalid value for {target}: {value}")
            raise e
        except AttributeError as e:
            print(f"Invalid target parameter: {target}")
        except TypeError as e:
            print(f"Invalid type for {target}: {value}")
        except Exception as e:
            print(f"Unexpected error: {e}")


class HistoryParam(BaseModel):
    user_name: str
    sample_name: str
    measure_type: str
    option: str


class ReferHistoryParam(BaseModel):
    operator: str
    material: str
    sample: str
    measure_type: str